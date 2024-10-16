from django.shortcuts import render
from django.http import HttpResponse
from .models import Soura
from .models import Part
from .models import Year
from .models import Student
from .models import Grade
from registration.models import NewStudent
from openpyxl import Workbook, load_workbook

# Create your views here.

def importData(request):
    # Fetch data from the database
    new_students = NewStudent.objects.all()
    students = Student.objects.all()
    items = []
    # for new_student in new_students:
    #     item = [] 
    #     if new_student.national_id:
    #         try:
    #             student = Student.objects.get(national_id= new_student.national_id)
    #             item.append(student.name)
    #             item.append(student.national_id)
    #             item.append(new_student.grade)
    #             item.append(len(items))
    #             items.append(item)
    #             grade = Grade()
    #             grade.grade = new_student.grade
    #             grade.student = student
    #             grade.year = Year.objects.get(year= '2024')
    #             grade.part = new_student.part
    #             grade.soura = new_student.soura
    #             grade.from_baqra = new_student.from_baqra
    #             # grade.save()
    #         except (Student.DoesNotExist):
    #             print('not found')
    
    # for new_student in new_students:
    #     item = [] 
    #     if new_student.national_id:
    #         try:
    #             student = Student.objects.get(national_id= new_student.national_id)
                # item.append(student.name)
                # item.append(student.national_id)
                # item.append(new_student.grade)
                # item.append(len(items))
                # items.append(item)
                # student.last_part = new_student.part
                # student.save()
            #     print('not found')
            # except (Student.DoesNotExist):
            #     student = Student()
            #     student.name = new_student.name
            #     student.national_id = new_student.national_id
            #     student.phone = new_student.phone
            #     student.next_amount = new_student.part
            #     student.last_part = new_student.part
            #     student.ahkam = "لا يوجد"
            #     student.save()
            #     item.append(new_student.name)
            #     item.append(new_student.national_id)
            #     item.append(new_student.grade)
            #     item.append(len(items))
            #     items.append(item)

    # count = 0
    # for student in students:
    #     item = [] 
    #     try:
    #         grades = Grade.objects.filter(student__national_id= student.national_id)
    #         for grade in grades:
    #             if int(grade.grade) > 0:
    #                 student.last_part = grade.part
    #                 student.last_soura = grade.soura
    #                 student.last_grade = grade.grade
    #                 student.save()
    #                 count += 1
    #                 break
    #     except (Student.DoesNotExist):
    #         print('not found')

    # item.append(count)
    # items.append(item)
    # for grade in Grade.objects.filter(student__national_id= "30407182200848"):
    #     item = [] 
    #     item.append(grade.year)
    #     items.append(item)

    # # Create a new Excel workbook
    # wb = Workbook()
    # ws = wb.active

    # # Add column headers
    # ws.append(['الاسم', 'الرقم القومي', 'التليفون', 'المقرر', 'الاحكام'])  # Replace with actual column names

    # # Add data rows
    # for row in data:
    #     ws.append([row.name, row.national_id, row.phone, row.next_amount.title, row.ahkam])  # Replace field1, field2, field3 with actual field names

    # # Create the response
    # response = HttpResponse(content_type='application/ms-excel')
    # response['Content-Disposition'] = 'attachment; filename="data.xlsx"'

    # # Save the workbook to the response
    # wb.save(response)

    # return response

    # grades = Grade.objects.all()
    # for grade in grades:
    #     if grade.student.national_id == '30310012209588' or grade.student.national_id == '30005012202485':
    #         item = [] 
    #         item.append(grade.student.name)
    #         item.append(grade.part.title)
    #         item.append(grade.soura.title)
    #         if (int(grade.part.number) <= 15 and int(grade.soura.number) >= 18) or (int(grade.part.number) > 15 and int(grade.soura.number) < 18):
    #             item.append('من سورة الناس الي سورة ' + grade.soura.title)
    #         else:
    #             item.append('من سورة الفاتحة الي سورة ' + grade.soura.title)
    #         items.append(item)

    # grades = Grade.objects.all()
    # for grade in grades:
    #     if grade.student.national_id == '31011032202535': # wrong national id
            # try:
            #     student = Student.objects.get(national_id= '31011032202525')
            #     grade.student = student
            #     grade.save()
            #     print('saved')
            # except (Student.DoesNotExist):
            #     print('not found')
            # student = grade.student
            # student.isFinished = True
            # student.save()
            # item = [] 
            # item.append(grade.student.name)
            # item.append(grade.student.phone)
            # item.append(grade.soura.title)
            # item.append(grade.grade)
            # item.append(grade.student.next_amount.title)
            # items.append(item)
    
    
    # excelFile = load_workbook('/Volumes/Data/Shabab Al-Barqy/Shabab Elkhier/Website/2024.xlsx')
    # ws = excelFile['sheet']

    # count = 0
    # for i in range(2, 1104):
    #     name = ws['A'+str(i)].value
    #     national_id = ws['C'+str(i)].value
    #     prize_time = ws['P'+str(i)].value
    #     phone = ws['D'+str(i)].value

    #     if phone is None:
    #         phone = "0"

    #     try:
    #         grade = int(ws['M'+str(i)].value)
    #     except:
    #         print(national_id)

    #     try:
    #         soura = Soura.objects.get(title= ws['B'+str(i)].value)
    #     except:
    #         print(national_id)

    #     try:
    #         part = Part.objects.get(number= int(ws['E'+str(i)].value))
    #     except:
    #         print(national_id)

    #     try:
    #         student = NewStudent()
    #         if national_id != None and (grade > 0):
    #             count += 1
    #             student.name = name
    #             student.national_id = national_id
    #             student.phone = phone
    #             student.part = part
    #             student.soura = soura
    #             student.prize_time = prize_time
    #             student.grade = grade
    #             student.save()
    #             item = []
    #             item.append(count)
    #             item.append(name)
    #             # item.append(national_id)
    #             # item.append(phone)
    #             # item.append(part)
    #             # item.append(Soura.objects.get(title= soura))
    #             # item.append(grade)
    #             # item.append(prize_time)
    #             items.append(item)
    #     except:
    #         print('not found')
    # print(count)
    return render(request, 'importData/importData.html', {'items': items})
